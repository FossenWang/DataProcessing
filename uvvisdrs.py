'''
处理紫外可见漫反射光谱的实验数据
'''
import os
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from scipy.stats import linregress
from openpyxl import Workbook

class UvvisDrsData:
    '''
    紫外可见漫反射光谱数据，一份数据包含了波长数组、反射率数组和数据名
    '''
    def __init__(self, wavelength_array, reflectance_array, name):
        self.wavelength_array = wavelength_array
        self.reflectance_array = reflectance_array
        self.name = name    #保存文件的绝对路径或相对路径
        self.hv = self.calculate_hv(wavelength_array)
        self.fr = self.calculate_fr()
        self.hvfr2 = self.calculate_hvfr2()
        self.hvfr12 = self.calculate_hvfr12()
        self.hvfr2_logi_fit = logistic_fit(self.hv, self.hvfr2)
        self.hvfr12_logi_fit = logistic_fit(self.hv, self.hvfr12)
        self.egd, self.kd, self.bd, self.rd = self.calculate_eg(self.hvfr2, self.hvfr2_logi_fit)
        self.egi, self.ki, self.bi, self.ri = self.calculate_eg(self.hvfr12, self.hvfr12_logi_fit)

    def calculate_hv(self, wavelength_array):
        '''
        hv=hc/λ=1240/λ
        用该式将波长换算成能量hv
        '''
        return 1240/wavelength_array

    def calculate_fr(self):
        '''
        Kubelka–Munk absorption function: F(R) = (1-R)^2/(2R)
        用上式计算出F(R)的值，其中R为reflectance
        '''
        return ((1-self.reflectance_array)**2)/(2*self.reflectance_array)

    def calculate_hvfr2(self):
        '''
        计算(hvF(R))^2
        '''
        return (self.hv*self.fr)**2

    def calculate_hvfr12(self):
        '''
        计算(hvF(R))^1/2
        '''
        return (self.hv*self.fr)**0.5

    def calculate_eg(self, y, y_fit=0, fp=0, a=25):
        '''
        y_fit经过logistic拟合的y，对y_fit数值微分，求得微分最大值，取得该点的引索imdy
        分别取该点周围n个(5<=n<=50)元素进行线性拟合，返回斜率k，截距b,相关系数r和imdy
        '''
        if not fp:
            dy=num_differ(self.hv, y_fit)
            imdy=np.where(dy==max(dy))[0][0]
        else:
            absdif = np.fabs(self.hv-fp)
            imdy=np.where(absdif==min(absdif))[0][0]
        for i in range(a,1,-1):
            if (imdy-i)>=0 and (imdy+i)<len(self.hv):
                k, b, r, p, s = linregress(self.hv[(imdy-i):(imdy+i)], y[(imdy-i):(imdy+i)])
                if r>0.99:
                    break
        eg = -b/k
        return [eg, k, b, r]

    def refit(self, n, fp=0, a=25):
        '''
        根据输入的点x0和左右元素范围a重新拟合曲线，n为（hvfr）^n中的的指数n
        '''
        if n == 2:
            self.egd, self.kd, self.bd, self.rd = self.calculate_eg(self.hvfr2, fp=fp, a=a)
            print('done')
        elif n == 0.5:
            self.egi, self.ki, self.bi, self.ri = self.calculate_eg(self.hvfr12, fp=fp, a=a)
            print('done')
        else:
            print('please enter n=2 or n = 0.5')
    
    def write_txt(self):
        '''
        将数据写入至txt文件,保存路径为self.name+'_result.txt'
        '''
        with open(self.name.split('.')[0]+'_result.txt', 'w') as txt:
            txt.write('samlpe:  '+self.name+'\n')
            txt.write('direct band gap: wevelength='+str(1240/self.egd)+'nm energy='+str(self.egd)+'eV\n')
            txt.write('\tlinear fit coefficient: slope='+str(self.kd)+' intercept='+str(self.bd)+' r='+str(self.rd))
            txt.write('\nindirect band gap: wevelength='+str(1240/self.egi)+'nm energy='+str(self.egi)+'eV\n')
            txt.write('\tlinear fit coefficient: slope='+str(self.ki)+' intercept='+str(self.bi)+' r='+str(self.ri))
            txt.write('\ndatas:\n')
            txt.write('wavelength\tR\thv\tF(R)\t(hvF(R))^2\t(hvF(R))^1/2\n')
            for data in zip(self.wavelength_array, self.reflectance_array, self.hv, self.fr, self.hvfr2, self.hvfr12):
                line = ''
                for each in data:
                    line = line + str(each) + '\t'
                txt.write(line[:-1] + '\n')

        print('已保存至  ', self.name.split('.')[0]+'_result.txt')

    def write_xlsx(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:F1')
        ws['A1'] = 'samlpe:  '+self.name
        ws.merge_cells('A2:F2')
        ws['A2'] = 'direct band gap: wevelength='+str(1240/self.egd)+'nm energy='+str(self.egd)+'eV'
        ws.merge_cells('A3:F3')
        ws['A3'] = 'linear fit coefficient: slope='+str(self.kd)+' intercept='+str(self.bd)+' r='+str(self.rd)
        ws.merge_cells('A4:F4')
        ws['A4'] = 'indirect band gap: wevelength='+str(1240/self.egi)+'nm energy='+str(self.egi)+'eV'
        ws.merge_cells('A5:F5')
        ws['A5'] = 'linear fit coefficient: slope='+str(self.ki)+' intercept='+str(self.bi)+' r='+str(self.ri)
        ws['A'+str(6)] = 'datas:'
        ws['A'+str(7)] = 'wavelength'
        ws['B'+str(7)] = 'R'
        ws['C'+str(7)] = 'hv'
        ws['D'+str(7)] = 'F(R)'
        ws['E'+str(7)] = '(hvF(R))^2'
        ws['F'+str(7)] = '(hvF(R))^1/2'
        i=7
        for data in zip(self.wavelength_array, self.reflectance_array, self.hv, self.fr, self.hvfr2, self.hvfr12):
            i+=1
            ws['A'+str(i)] = data[0]
            ws['B'+str(i)] = data[1]
            ws['C'+str(i)] = data[2]
            ws['D'+str(i)] = data[3]
            ws['E'+str(i)] = data[4]
            ws['F'+str(i)] = data[5]
        path = self.name.split('.')[0]+'_result.xlsx'
        try:
            wb.save(path)
            print('已保存至  ', path)
        except PermissionError as e:
            print(e, '\n无法保存文件，请确认该文件是否被其他程序打开')

    def draw_hvfr(self):
        '''
        传入UvvisDrsData实例
        绘制(hvF(R))^2-hv图和(hvF(R))^1/2-hv图
        '''
        fig = plt.figure()
        fig.set_size_inches(12,5.2)
        ax1 = fig.add_subplot(121)
        ax2 = fig.add_subplot(122)
        ax1.plot(self.hv, self.hvfr2, label='direct')
        ax1.plot([self.egd, (max(self.hvfr2)-self.bd)/self.kd], [0, max(self.hvfr2)])
        ax1.set_xlabel('hv (eV)')
        ax1.set_ylabel(r'$(hv F(R))^2$')
        handles, labels = ax1.get_legend_handles_labels()
        ax1.legend(handles, labels)
        ax1.text(self.egd-0.25, max(self.hvfr2)/2,
            (r'$\lambda$ = '+str(int(round(1240/self.egd)))+' nm\n'+r'$E_g$ = '+str(round(self.egd,2))+' eV'),
            horizontalalignment='right', verticalalignment='bottom')

        ax2.plot(self.hv, self.hvfr12, label='indirect')
        ax2.plot([self.egi, (max(self.hvfr12)-self.bi)/self.ki], [0, max(self.hvfr12)])
        ax2.set_xlabel('hv (eV)')
        ax2.set_ylabel(r'$(hv F(R))^\frac{1}{2}$')
        handles, labels = ax2.get_legend_handles_labels()
        ax2.legend(handles, labels)
        ax2.text(self.egi-0.25, max(self.hvfr12)/2,
            (r'$\lambda$ = '+str(int(round(1240/self.egi)))+' nm\n'+r'$E_g$ = '+str(round(self.egi,2))+' eV'),
            horizontalalignment='right', verticalalignment='bottom')
        return fig

def logistic_fit(x, y):
    '''
    拟合logistic函数,P(t)=KP_0exp(rt)/(K+P_0(exp(rt)-1)),P_0,为初值，K为终值
    '''
    def func(z, t, r):
        return (k*p*np.exp(r*z+t))/(k+p*(np.exp(r*z+t)-1))

    k = max(y)
    p = y[-1]
    i = np.where(y==k)[0][0]

    popt, pcov = curve_fit(func, x[i:], y[i:])
    return func(x, *popt)

def num_differ(x, y):
    '''
    numerical differentiation 数值微分
    参数为numpy.ndarray对象
    '''
    x1=np.append(x[1:], x[-1])
    x2=np.append(x[0], x[:-1])
    y1=np.append(y[1:], y[-1])
    y2=np.append(y[0], y[:-1])
    return (y1-y2)/(x1-x2)

def read_raw(file):
    '''
    读取存有DRS数据的txt文件
    支持读取以下型号的仪器产生的数据：
    UV-Visible diffuse reflectance spectroscopy UV-2550PC (Shimadzu Corporation, Japan)
    '''
    with open(file, 'r') as txt:
        txtline = '1'
        while txtline != '':
            txtline = txt.readline()
            if txtline == '"波长(nm)","T%"\n':
                break
        wavelength, reflectance_array = [], []
        txtline = txt.readline()
        while txtline != '':
            txtline = txtline.strip('\n')
            wavelength.append(float(txtline.split(',')[0]))
            reflectance_array.append(float(txtline.split(',')[1]))
            txtline = txt.readline()

    return UvvisDrsData(np.array(wavelength), np.array(reflectance_array)/100, file)
